import openpyxl

wb = openpyxl.load_workbook('/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/SIS_IM_v1.1.3_20250319_1352.xlsx', data_only=True)

with open('/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/excel_analysis.txt', 'w', encoding='utf-8') as out:
    for name in wb.sheetnames:
        ws = wb[name]
        out.write(f"=== SHEET: {name} (max_row={ws.max_row}, max_col={ws.max_column}) ===\n")
        headers = [str(cell.value or '').strip() for cell in ws[1]]
        out.write(f"Headers: {headers}\n\n")
        for r in range(2, min(5, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            out.write(f"Row {r-1}: {vals}\n")
        out.write("\n" + "="*50 + "\n\n")

print("Wrote analysis to excel_analysis.txt")
