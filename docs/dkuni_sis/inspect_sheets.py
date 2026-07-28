import openpyxl

wb = openpyxl.load_workbook('SIS_IM_v1.1.3_20250319_1352.xlsx', data_only=True)
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"=== Sheet: {sheetname} ===")
    headers = [str(cell.value or '').strip() for cell in ws[1]]
    print("Headers:", headers)
    for r in range(2, min(7, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        print(f"Row {r-1}: {row_vals}")
    print()
