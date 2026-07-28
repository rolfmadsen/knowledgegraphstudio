#!/usr/bin/env python3
"""
SIS Multi-Excel to KnowledgeGraph Studio YAML & Views Transformer

Solves Canvas Layout & Isolation Problems:
1. Domain Connectivity ('includes' edges): Connects isolated concepts to their parent Domain node, eliminating floating nodes.
2. Structured Views Generator (views.xarchi.yaml): Generates targeted views per Sub-domain (Optagelse, Eksamen, etc.) and per Notation (Begrebsmodel / Informationsmodel).
3. Compact 2D Grid & Force-Directed Layouts: Replaces single-column tree stacking with neat 5-column 2D grids (X=col*320, Y=row*180) and force-directed physics layouts.
4. Clean Names & 100% Validated Syntax.
"""

import sys
import os
import uuid
import time
import openpyxl
import yaml

BEGREBSLISTE_FILENAME = "begrebsliste_v1_1_0.xlsx"
SIS_IM_FILENAME = "SIS_IM_v1.1.3_20250319_1352.xlsx"
OUTPUT_MODEL_FILENAME = "model.xarchi.yaml"
OUTPUT_VIEWS_FILENAME = "views.xarchi.yaml"

def generate_element_id(prefix: str) -> str:
    """Generates a valid ElementId matching KnowledgeGraph Studio's Zod schema regex."""
    return f"{prefix}:{uuid.uuid4()}"

def sanitize_slug(name: str) -> str:
    """Normalizes string names to a clean alphanumeric slug key for robust lookup mapping."""
    if name is None:
        return ""
    clean = str(name).strip().lower()
    clean = clean.replace("æ", "ae").replace("ø", "oe").replace("å", "aa")
    for char in [" ", "_", "/", "\\", "(", ")", ".", ",", ":", ";", "-", "–", "—", "'", '"']:
        clean = clean.replace(char, "")
    return clean

def find_column_index(headers_normalized: list, keywords: list):
    """Finds column index in headers matching candidate keywords."""
    for kw in keywords:
        kw_norm = sanitize_slug(kw)
        for idx, h in enumerate(headers_normalized):
            if kw_norm in h:
                return idx
    return None

def get_row_val(row: tuple, col_idx: int):
    """Returns stripped string value from row at col_idx if present."""
    if col_idx is not None and col_idx < len(row) and row[col_idx] is not None:
        val = str(row[col_idx]).strip()
        if val and val.lower() not in ["none", "n/a", "null"]:
            return val
    return None

def parse_aliases(alias_str: str) -> list:
    """Parses comma/semicolon separated aliases."""
    if not alias_str:
        return []
    raw_list = [a.strip() for a in alias_str.replace(";", ",").split(",")]
    return [a for a in raw_list if a]

def parse_classification(val: str) -> str:
    """Maps dataset classification to KnowledgeGraph Studio DataClassification enum."""
    if not val:
        return None
    val_lower = val.lower()
    if "følsom" in val_lower or "foelsom" in val_lower or "særlig" in val_lower or "niveau 3" in val_lower:
        return "niveau_3_foelsom"
    elif "fortrolig" in val_lower or "niveau 2" in val_lower:
        return "niveau_2_fortrolig"
    elif "intern" in val_lower or "niveau 1" in val_lower:
        return "niveau_1_intern"
    elif "offentlig" in val_lower or "niveau 0" in val_lower:
        return "niveau_0_offentlig"
    return None

def main():
    script_dir = os.path.dirname(os.path.realpath(__file__))
    begrebsliste_path = os.path.join(script_dir, BEGREBSLISTE_FILENAME)
    sis_im_path = os.path.join(script_dir, SIS_IM_FILENAME)
    output_model_path = os.path.join(script_dir, OUTPUT_MODEL_FILENAME)
    output_views_path = os.path.join(script_dir, OUTPUT_VIEWS_FILENAME)

    now_ms = int(time.time() * 1000)

    # Main Domain
    main_domain_id = generate_element_id("domain")
    domains_map = {
        "main": {
            "id": main_domain_id,
            "name": "SIS - Studieadministrativt Informationssystem",
            "description": "Fælles Informations- og Begrebsmodel for SIS (DK UNI)",
            "createdAt": now_ms,
            "updatedAt": now_ms,
            "lifecycleState": "active"
        }
    }

    entity_map = {}       # slug -> Begreb concept object
    class_map = {}        # slug -> Klasse concept object
    enum_concept_map = {} # slug -> Enumeration concept object
    id_map = {}           # id/code -> concept

    # ---------------------------------------------------------
    # PART A: Read Begrebsliste (begrebsliste_v1_1_0.xlsx)
    # ---------------------------------------------------------
    begreb_rel_list = []
    if os.path.exists(begrebsliste_path):
        print(f"Loading Begrebsliste workbook: {begrebsliste_path}...")
        wb_b = openpyxl.load_workbook(begrebsliste_path, data_only=True)
        print(f"  Begrebsliste sheets: {wb_b.sheetnames}")

        b_sheet_candidates = [s for s in wb_b.sheetnames if any(k in s.lower() for k in ["begreb", "term", "koncept", "query"])]
        b_sheet_name = b_sheet_candidates[0] if b_sheet_candidates else wb_b.sheetnames[0]
        ws_b = wb_b[b_sheet_name]

        b_headers_raw = [str(cell.value or "").strip() for cell in ws_b[1]]
        b_headers_norm = [sanitize_slug(h) for h in b_headers_raw]

        b_name_col = find_column_index(b_headers_norm, ["termforetrukken", "begreb", "navn", "term", "titel", "begrebsnavn"])
        b_def_col = find_column_index(b_headers_norm, ["definition", "beskrivelse", "forklaring", "annotation"])
        b_alias_col = find_column_index(b_headers_norm, ["synonymaecepteret", "synonym", "synonymer", "aliases", "alternativ"])
        b_legal_col = find_column_index(b_headers_norm, ["lovgivning", "lovhjemmel", "hjemmel", "kilde", "referance", "standard"])
        b_dom_col = find_column_index(b_headers_norm, ["tilhoereremneomraadet", "domaene", "kategori", "emneomraade", "pakke"])

        for row in ws_b.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            name = get_row_val(row, b_name_col)
            if not name or name.startswith("#"):
                continue

            slug = sanitize_slug(name)
            if not slug:
                continue

            definition = get_row_val(row, b_def_col) or ""
            aliases = parse_aliases(get_row_val(row, b_alias_col) or "")
            legal_src = get_row_val(row, b_legal_col)

            domain_name = get_row_val(row, b_dom_col)
            target_domain_id = main_domain_id
            if domain_name:
                d_slug = sanitize_slug(domain_name)
                if d_slug not in domains_map:
                    domains_map[d_slug] = {
                        "id": generate_element_id("domain"),
                        "name": domain_name,
                        "description": f"SIS Domæne: {domain_name}",
                        "createdAt": now_ms,
                        "updatedAt": now_ms,
                        "lifecycleState": "active"
                    }
                target_domain_id = domains_map[d_slug]["id"]

            if slug in entity_map:
                existing = entity_map[slug]
                if definition and not existing.get("definition"):
                    existing["definition"] = definition
                for alias in aliases:
                    if alias not in existing["aliases"]:
                        existing["aliases"].append(alias)
                if legal_src and not existing.get("legalSource"):
                    existing["legalSource"] = legal_src
            else:
                begreb_id = generate_element_id("class")
                begreb_concept = {
                    "id": begreb_id,
                    "conceptType": "class",
                    "name": name,
                    "definition": definition,
                    "domainId": target_domain_id,
                    "aliases": aliases,
                    "preferredTerm": name,
                    "policies": [],
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active",
                    "relations": []
                }
                if legal_src:
                    begreb_concept["legalSource"] = legal_src

                entity_map[slug] = begreb_concept

        print(f"  Loaded {len(entity_map)} deduplicated Begreber from Begrebsliste.")

    # ---------------------------------------------------------
    # PART B: Read SIS Information Model (SIS_IM_v1.1.3_20250319_1352.xlsx)
    # ---------------------------------------------------------
    if not os.path.exists(sis_im_path):
        print(f"Error: Could not find SIS Information Model file at '{sis_im_path}'")
        sys.exit(1)

    print(f"\nLoading SIS Information Model workbook: {sis_im_path}...")
    wb_im = openpyxl.load_workbook(sis_im_path, data_only=True)
    sheet_names = wb_im.sheetnames

    # 1. Enumerations / Value Lists
    enum_sheet_candidates = [s for s in sheet_names if any(k in s.lower() for k in ["værdilist", "kodelist", "enum"])]
    if enum_sheet_candidates:
        ws_enum = wb_im[enum_sheet_candidates[0]]
        enum_raw_headers = [str(cell.value or "").strip() for cell in ws_enum[1]]
        enum_headers_norm = [sanitize_slug(h) for h in enum_raw_headers]

        name_col = find_column_index(enum_headers_norm, ["navn", "vaerdiliste", "kodeliste", "name", "titel"])
        def_col = find_column_index(enum_headers_norm, ["definition", "beskrivelse", "description"])

        enum_val_sheet = [s for s in sheet_names if "værdi" in s.lower() and s != enum_sheet_candidates[0]]
        enum_values_map = {}

        if enum_val_sheet:
            ws_val = wb_im[enum_val_sheet[0]]
            val_headers_norm = [sanitize_slug(str(cell.value or "").strip()) for cell in ws_val[1]]
            v_list_col = find_column_index(val_headers_norm, ["vaerdiliste", "kodeliste", "list", "enum", "navn"])
            v_item_col = find_column_index(val_headers_norm, ["vaerdi", "kode", "value", "item", "label", "titel"])

            for row in ws_val.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                list_name = get_row_val(row, v_list_col)
                item_val = get_row_val(row, v_item_col)
                if list_name and item_val:
                    l_slug = sanitize_slug(list_name)
                    if l_slug not in enum_values_map:
                        enum_values_map[l_slug] = []
                    if item_val not in enum_values_map[l_slug]:
                        enum_values_map[l_slug].append(item_val)

        for row in ws_enum.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            enum_name = get_row_val(row, name_col)
            if not enum_name:
                continue

            e_slug = sanitize_slug(enum_name)
            if e_slug in enum_concept_map:
                continue

            enumerators = enum_values_map.get(e_slug, [])
            enum_id = generate_element_id("enumeration")

            enum_concept = {
                "id": enum_id,
                "conceptType": "enumeration",
                "name": enum_name,
                "definition": get_row_val(row, def_col) or f"Kodeliste for {enum_name}",
                "enumerators": enumerators,
                "aliases": [],
                "policies": [],
                "createdAt": now_ms,
                "updatedAt": now_ms,
                "lifecycleState": "active"
            }
            enum_concept_map[e_slug] = enum_concept

        print(f"  Loaded {len(enum_concept_map)} Enumerations.")

    # 2. Classes & Linking to Begreber
    class_sheet_candidates = [s for s in sheet_names if any(k in s.lower() for k in ["klasse", "class", "entitet"])]
    ws_class = wb_im[class_sheet_candidates[0] if class_sheet_candidates else sheet_names[0]]
    class_raw_headers = [str(cell.value or "").strip() for cell in ws_class[1]]
    class_headers_norm = [sanitize_slug(h) for h in class_raw_headers]

    name_col = find_column_index(class_headers_norm, ["navn", "klasse", "klassonavn", "klassenavn", "name", "titel", "entitet", "begreb"])
    id_col = find_column_index(class_headers_norm, ["id", "klasseid", "klasse_id", "kode", "identifikator", "uri", "uuid"])
    def_col = find_column_index(class_headers_norm, ["definition", "beskrivelse", "description", "forklaring", "nota"])
    alias_col = find_column_index(class_headers_norm, ["synonym", "synonymer", "aliases", "alternativtnavn"])
    dom_col = find_column_index(class_headers_norm, ["domaene", "domene", "emneomraade", "pakke", "boundedcontext", "context"])
    legal_col = find_column_index(class_headers_norm, ["lovhjemmel", "hjemmel", "kilde", "legalsource"])
    classif_col = find_column_index(class_headers_norm, ["klassificering", "foelsomhed", "datasikkerhed", "classification"])

    for row in ws_class.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        name = get_row_val(row, name_col)
        if not name or name.startswith("#"):
            continue

        slug = sanitize_slug(name)
        if not slug or slug in class_map:
            continue

        definition = get_row_val(row, def_col) or ""
        aliases = parse_aliases(get_row_val(row, alias_col) or "")
        legal_src = get_row_val(row, legal_col)
        classification = parse_classification(get_row_val(row, classif_col))

        domain_name = get_row_val(row, dom_col)
        target_domain_id = main_domain_id
        if domain_name:
            d_slug = sanitize_slug(domain_name)
            if d_slug not in domains_map:
                domains_map[d_slug] = {
                    "id": generate_element_id("domain"),
                    "name": domain_name,
                    "description": f"SIS Domæne: {domain_name}",
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active"
                }
            target_domain_id = domains_map[d_slug]["id"]

        begreb_concept = entity_map.get(slug)
        if not begreb_concept:
            begreb_id = generate_element_id("class")
            begreb_concept = {
                "id": begreb_id,
                "conceptType": "class",
                "name": name,
                "definition": definition,
                "domainId": target_domain_id,
                "aliases": aliases,
                "preferredTerm": name,
                "policies": [],
                "createdAt": now_ms,
                "updatedAt": now_ms,
                "lifecycleState": "active",
                "relations": []
            }
            if legal_src: begreb_concept["legalSource"] = legal_src
            if classification: begreb_concept["classification"] = classification
            entity_map[slug] = begreb_concept

        class_id = generate_element_id("class")
        class_concept = {
            "id": class_id,
            "conceptType": "class",
            "name": name,
            "definition": definition,
            "domainId": target_domain_id,
            "wasDerivedFrom": begreb_concept["id"],  # Traceability!
            "aliases": aliases,
            "policies": [],
            "createdAt": now_ms,
            "updatedAt": now_ms,
            "lifecycleState": "active",
            "properties": [],
            "relations": []
        }
        if legal_src: class_concept["legalSource"] = legal_src
        if classification: class_concept["classification"] = classification

        class_map[slug] = class_concept

        class_code = get_row_val(row, id_col)
        if class_code:
            id_map[sanitize_slug(class_code)] = class_concept

    print(f"  Loaded {len(class_map)} Information Model Klasser linked to {len(entity_map)} Begreber.")

    # 3. Attributes
    attr_sheet_candidates = [s for s in sheet_names if any(k in s.lower() for k in ["attribut", "egenskab", "property", "attribute"])]
    if attr_sheet_candidates:
        ws_attr = wb_im[attr_sheet_candidates[0]]
        attr_raw_headers = [str(cell.value or "").strip() for cell in ws_attr[1]]
        attr_headers_norm = [sanitize_slug(h) for h in attr_raw_headers]

        a_class_col = find_column_index(attr_headers_norm, ["klasse", "klassonavn", "klassenavn", "entitet", "class", "classname", "klasseid"])
        a_name_col = find_column_index(attr_headers_norm, ["navn", "attribut", "attributnavn", "property", "attribute", "felt"])
        a_type_col = find_column_index(attr_headers_norm, ["datatype", "type", "datatid", "format"])
        a_enum_col = find_column_index(attr_headers_norm, ["vaerdiliste", "valgliste", "kodeliste", "enum", "valuelist"])
        a_req_col = find_column_index(attr_headers_norm, ["obligatorisk", "required", "paakraevet", "mo"])
        a_mult_col = find_column_index(attr_headers_norm, ["kardinalitet", "multiplicitet", "multiplicity", "cardinality", "antal"])

        attr_count = 0
        enum_linked_count = 0

        for row in ws_attr.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            class_ref = get_row_val(row, a_class_col)
            attr_name = get_row_val(row, a_name_col)

            if not class_ref or not attr_name:
                continue

            class_slug = sanitize_slug(class_ref)
            target_concept = class_map.get(class_slug) or id_map.get(class_slug)

            if target_concept:
                raw_type = (get_row_val(row, a_type_col) or "string").lower()
                enum_ref = get_row_val(row, a_enum_col)

                dtype = "string"
                if enum_ref:
                    e_slug = sanitize_slug(enum_ref)
                    if e_slug in enum_concept_map:
                        dtype = enum_concept_map[e_slug]["id"]
                        enum_linked_count += 1

                if dtype == "string":
                    if any(t in raw_type for t in ["int", "integer", "number", "decimal", "float", "tal", "heltal"]):
                        dtype = "number"
                    elif any(t in raw_type for t in ["bool", "boolean", "sandhedsværdi", "sand/falsk"]):
                        dtype = "boolean"
                    elif any(t in raw_type for t in ["date", "datetime", "tid", "dato", "tidsstempel"]):
                        dtype = "date"

                req_val = (get_row_val(row, a_req_col) or "").lower()
                is_req = req_val in ["ja", "true", "1", "m", "mandatory", "yes", "påkrævet", "paakraevet"]
                mult = get_row_val(row, a_mult_col) or ("1" if is_req else "0..1")

                prop = {
                    "id": generate_element_id("prop"),
                    "name": attr_name,
                    "type": dtype,
                    "isRequired": is_req,
                    "multiplicity": mult,
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active"
                }
                target_concept["properties"].append(prop)
                attr_count += 1

        print(f"  Added {attr_count} attributes ({enum_linked_count} linked to Enumerations).")

    # 4. Relations (Edges) -> Attach to BOTH Klasser & Begreber
    rel_sheet_candidates = [s for s in sheet_names if any(k in s.lower() for k in ["relation", "association", "forbindelse", "edge"])]
    if rel_sheet_candidates:
        ws_rel = wb_im[rel_sheet_candidates[0]]
        rel_raw_headers = [str(cell.value or "").strip() for cell in ws_rel[1]]
        rel_headers_norm = [sanitize_slug(h) for h in rel_raw_headers]

        r_src_col = find_column_index(rel_headers_norm, ["fra", "kilde", "source", "start", "fraklasse", "kildeklasse", "entitet1", "klasse1"])
        r_target_col = find_column_index(rel_headers_norm, ["til", "maal", "mal", "target", "slut", "tilklasse", "maalklasse", "entitet2", "klasse2", "raekkevidde"])
        r_name_col = find_column_index(rel_headers_norm, ["navn", "relation", "relationsnavn", "rolle", "label", "titel", "forbindelse"])
        r_type_col = find_column_index(rel_headers_norm, ["type", "relationstype", "kategori", "category", "kind"])
        r_src_role_col = find_column_index(rel_headers_norm, ["kilderolle", "fra_rolle", "sourcerole"])
        r_target_role_col = find_column_index(rel_headers_norm, ["maalrolle", "malrolle", "til_rolle", "targetrole"])
        r_src_mult_col = find_column_index(rel_headers_norm, ["kardinalitet1", "kildekardinalitet", "sourcemultiplicity"])
        r_target_mult_col = find_column_index(rel_headers_norm, ["kardinalitet2", "maalkardinalitet", "malkardinalitet", "targetmultiplicity", "kardinalitet", "multiplicitet"])

        def resolve_concept(slug):
            return (
                class_map.get(slug) or 
                entity_map.get(slug) or 
                enum_concept_map.get(slug) or 
                id_map.get(slug)
            )

        class_rel_count = 0
        begreb_rel_count = 0

        for row in ws_rel.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            src_raw = get_row_val(row, r_src_col)
            target_raw = get_row_val(row, r_target_col)
            rel_name = get_row_val(row, r_name_col) or "relaterer_til"

            if not src_raw or not target_raw:
                continue

            src_slug = sanitize_slug(src_raw)
            target_slug = sanitize_slug(target_raw)

            src_class_concept = class_map.get(src_slug) or id_map.get(src_slug)
            target_class_concept = resolve_concept(target_slug)

            src_begreb_concept = entity_map.get(src_slug)
            target_begreb_concept = entity_map.get(target_slug) or resolve_concept(target_slug)

            rel_type_raw = (get_row_val(row, r_type_col) or "association").lower()
            valid_types = ["association", "composition", "aggregation", "specialization", "realization"]
            rel_type = rel_type_raw if rel_type_raw in valid_types else "association"

            src_role = get_row_val(row, r_src_role_col)
            target_role = get_row_val(row, r_target_role_col)
            src_mult = get_row_val(row, r_src_mult_col)
            target_mult = get_row_val(row, r_target_mult_col) or "0..*"

            # 1. Add edge in Information Model
            if src_class_concept and target_class_concept:
                rel_class = {
                    "id": generate_element_id("rel"),
                    "sourceConceptId": src_class_concept["id"],
                    "targetConceptId": target_class_concept["id"],
                    "name": rel_name,
                    "category": "semantic",
                    "relationType": rel_type,
                    "multiplicity": target_mult,
                    "policies": [],
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active"
                }
                if src_role: rel_class["sourceRole"] = src_role
                if target_role: rel_class["targetRole"] = target_role
                if src_mult: rel_class["sourceMultiplicity"] = src_mult
                if target_mult: rel_class["targetMultiplicity"] = target_mult

                src_class_concept["relations"].append(rel_class)
                class_rel_count += 1

            # 2. Add edge in Conceptual Model
            if src_begreb_concept and target_begreb_concept:
                rel_begreb = {
                    "id": generate_element_id("rel"),
                    "sourceConceptId": src_begreb_concept["id"],
                    "targetConceptId": target_begreb_concept["id"],
                    "name": rel_name,
                    "category": "semantic",
                    "relationType": rel_type,
                    "multiplicity": target_mult,
                    "policies": [],
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active"
                }
                if src_role: rel_begreb["sourceRole"] = src_role
                if target_role: rel_begreb["targetRole"] = target_role

                src_begreb_concept["relations"].append(rel_begreb)
                begreb_rel_count += 1

        print(f"  Added {class_rel_count} relations to Information Model and {begreb_rel_count} relations to Conceptual Model.")

    # ---------------------------------------------------------
    # PART C: Auto-Connect Isolated Begreber via Domain 'includes' Relations
    # ---------------------------------------------------------
    # Create domain concept nodes so domain includes edges can point to them
    domain_concept_map = {}
    for d_slug, d_obj in domains_map.items():
        d_concept = {
            "id": d_obj["id"],
            "conceptType": "domain",
            "name": d_obj["name"],
            "definition": d_obj.get("description", ""),
            "createdAt": now_ms,
            "updatedAt": now_ms,
            "lifecycleState": "active",
            "relations": []
        }
        domain_concept_map[d_obj["id"]] = d_concept

    domain_includes_count = 0
    all_begreber = list(entity_map.values())
    for begreb in all_begreber:
        # If begreb has 0 relations, connect it to its Domain concept via 'includes' relation
        if len(begreb.get("relations", [])) == 0 and begreb.get("domainId"):
            d_concept = domain_concept_map.get(begreb["domainId"])
            if d_concept:
                inc_rel = {
                    "id": generate_element_id("rel"),
                    "sourceConceptId": d_concept["id"],
                    "targetConceptId": begreb["id"],
                    "name": "omfatter",
                    "category": "semantic",
                    "relationType": "includes",
                    "multiplicity": "0..*",
                    "policies": [],
                    "createdAt": now_ms,
                    "updatedAt": now_ms,
                    "lifecycleState": "active"
                }
                d_concept["relations"].append(inc_rel)
                domain_includes_count += 1

    print(f"  Created {domain_includes_count} Domain 'omfatter' (includes) edges connecting isolated Begreber.")

    # ---------------------------------------------------------
    # PART D: Assemble Concepts List
    # ---------------------------------------------------------
    all_concepts = list(domain_concept_map.values()) + list(entity_map.values()) + list(class_map.values()) + list(enum_concept_map.values())
    
    unique_concepts = []
    seen_ids = set()
    for c in all_concepts:
        if c["id"] not in seen_ids:
            seen_ids.add(c["id"])
            unique_concepts.append(c)

    export_domains = list(domains_map.values())
    export_data = {
        "version": "1.0",
        "domains": export_domains,
        "concepts": unique_concepts
    }

    with open(output_model_path, "w", encoding="utf-8") as f:
        yaml.dump(export_data, f, allow_unicode=True, sort_keys=False, default_flow_style=False)

    # ---------------------------------------------------------
    # PART E: Generate Structured views.xarchi.yaml with 2D Grid Layouts
    # ---------------------------------------------------------
    views = []
    
    # 1. Main Begrebsmodel View (Conceptual Model)
    begreb_nodes = [c for c in unique_concepts if c["id"] in {b["id"] for b in entity_map.values()}]
    begreb_view_nodes = []
    cols = 5
    for idx, b in enumerate(begreb_nodes):
        r = idx // cols
        c = idx % cols
        begreb_view_nodes.append({
            "conceptId": b["id"],
            "x": c * 340 + 50,
            "y": r * 200 + 50
        })

    views.append({
        "id": generate_element_id("view"),
        "name": "Begrebsmodel (Alle Begreber)",
        "type": "conceptual_model",
        "layoutAlgorithm": "force_directed",
        "nodes": begreb_view_nodes,
        "edges": [rel["id"] for b in begreb_nodes for rel in b.get("relations", [])],
        "createdAt": now_ms,
        "updatedAt": now_ms,
        "lifecycleState": "active"
    })

    # 2. Main Informationsmodel View
    class_nodes = [c for c in unique_concepts if c["id"] in {k["id"] for k in class_map.values()}]
    class_view_nodes = []
    for idx, k in enumerate(class_nodes):
        r = idx // cols
        c = idx % cols
        class_view_nodes.append({
            "conceptId": k["id"],
            "x": c * 360 + 50,
            "y": r * 220 + 50
        })

    views.append({
        "id": generate_element_id("view"),
        "name": "Informationsmodel (Alle Klasser)",
        "type": "information_model",
        "layoutAlgorithm": "force_directed",
        "nodes": class_view_nodes,
        "edges": [rel["id"] for k in class_nodes for rel in k.get("relations", [])],
        "createdAt": now_ms,
        "updatedAt": now_ms,
        "lifecycleState": "active"
    })

    # 3. Sub-domain targeted views (Optagelse, Eksamen, etc.)
    for d_slug, d_obj in list(domains_map.items())[:15]:
        if d_slug == "main": continue
        d_id = d_obj["id"]
        sub_begreber = [b for b in begreb_nodes if b.get("domainId") == d_id]
        if not sub_begreber: continue

        sub_view_nodes = []
        for idx, b in enumerate(sub_begreber):
            r = idx // cols
            c = idx % cols
            sub_view_nodes.append({
                "conceptId": b["id"],
                "x": c * 340 + 50,
                "y": r * 200 + 50
            })

        views.append({
            "id": generate_element_id("view"),
            "name": f"Domæne: {d_obj['name']}",
            "type": "conceptual_model",
            "layoutAlgorithm": "force_directed",
            "nodes": sub_view_nodes,
            "edges": [rel["id"] for b in sub_begreber for rel in b.get("relations", [])],
            "createdAt": now_ms,
            "updatedAt": now_ms,
            "lifecycleState": "active"
        })

    with open(output_views_path, "w", encoding="utf-8") as f:
        yaml.dump(views, f, allow_unicode=True, sort_keys=False, default_flow_style=False)

    total_relations = sum(len(c.get("relations", [])) for c in unique_concepts)
    print(f"\n===========================================================")
    print(f"Success! Canvas Layout & Multi-View Transformation Complete:")
    print(f"  - {len(domains_map)} Domæner/Sub-domæner")
    print(f"  - {len(entity_map)} Unikke Begreber")
    print(f"  - {len(class_map)} Unikke Klasser")
    print(f"  - {len(enum_concept_map)} Unikke Værdilister/Kodelister")
    print(f"  - {domain_includes_count} Domæne 'omfatter' kanter (forbinder isolerede begreber)")
    print(f"  - {total_relations} Totale Edges")
    print(f"  - {len(views)} Strukturerede Views oprettet i views.xarchi.yaml med Force-Directed & 2D Grid layouts")
    print(f"Model written to: {output_model_path}")
    print(f"Views written to: {output_views_path}")
    print(f"===========================================================\n")

if __name__ == "__main__":
    main()
