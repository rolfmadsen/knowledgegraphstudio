import openpyxl

file_path = '/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/begrebsliste_v1_1_0.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

with open('/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/begrebsliste_info.txt', 'w', encoding='utf-8') as out:
    out.write(f"SHEET NAMES: {wb.sheetnames}\n\n")
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [str(cell.value or '').strip() for cell in ws[1]]
        out.write(f"=== SHEET: {name} (max_row={ws.max_row}, max_col={ws.max_column}) ===\n")
        out.write(f"Headers: {headers}\n")
        for r in range(2, min(6, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            out.write(f"Row {r-1}: {vals}\n")
        out.write("\n" + "="*50 + "\n\n")

print("Wrote analysis to begrebsliste_info.txt")
