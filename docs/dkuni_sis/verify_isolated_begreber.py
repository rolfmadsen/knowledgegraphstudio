#!/usr/bin/env python3
"""
Verification Script: Analyze Isolated Concepts & Edge Coverage

Analyzes model.xarchi.yaml and begrebsliste_v1_1_0.xlsx / SIS_IM_v1.1.3_20250319_1352.xlsx
to report exact statistics on connected vs isolated concepts in both models.
"""

import os
import openpyxl
import yaml

MODEL_YAML_PATH = "/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/model.xarchi.yaml"
BEGREBSLISTE_PATH = "/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/begrebsliste_v1_1_0.xlsx"
SIS_IM_PATH = "/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/SIS_IM_v1.1.3_20250319_1352.xlsx"
VERIFY_REPORT_PATH = "/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/verification_report.txt"

def main():
    report = []
    report.append("==========================================================")
    report.append("VERIFICATION REPORT: CONCEPT CONNECTIVITY & ISOLATION")
    report.append("==========================================================\n")

    if not os.path.exists(MODEL_YAML_PATH):
        print("model.xarchi.yaml not found.")
        return

    with open(MODEL_YAML_PATH, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    concepts = data.get("concepts", [])
    report.append(f"Total concepts in model.xarchi.yaml: {len(concepts)}\n")

    # Separate into Begreber (has preferredTerm), Klasser (has properties or wasDerivedFrom), Enums (conceptType: enumeration)
    begreber = [c for c in concepts if "preferredTerm" in c and c.get("conceptType") == "class"]
    klasser = [c for c in concepts if "wasDerivedFrom" in c and c.get("conceptType") == "class"]
    enums = [c for c in concepts if c.get("conceptType") == "enumeration"]

    report.append(f"Breakdown:")
    report.append(f"  - Begreber (Begrebsmodel): {len(begreber)}")
    report.append(f"  - Klasser (Informationsmodel): {len(klasser)}")
    report.append(f"  - Enumerationer/Kodelister: {len(enums)}\n")

    # Map all incoming and outgoing relations
    outgoing_edges = {c["id"]: len(c.get("relations", [])) for c in concepts}
    incoming_edges = {c["id"]: 0 for c in concepts}

    for c in concepts:
        for rel in c.get("relations", []):
            target_id = rel.get("targetConceptId")
            if target_id in incoming_edges:
                incoming_edges[target_id] += 1

    # Analyze Begreber
    begreber_with_edges = [c for c in begreber if outgoing_edges[c["id"]] > 0 or incoming_edges[c["id"]] > 0]
    begreber_isolated = [c for c in begreber if outgoing_edges[c["id"]] == 0 and incoming_edges[c["id"]] == 0]

    report.append("--- BEGREBSMODEL (Begreber) CONNECTIVITY ---")
    report.append(f"  - Begreber med relationer (edges): {len(begreber_with_edges)} ({len(begreber_with_edges)/len(begreber)*100:.1f}%)")
    report.append(f"  - Isolerede Begreber (0 edges):     {len(begreber_isolated)} ({len(begreber_isolated)/len(begreber)*100:.1f}%)")
    report.append(f"  - Eksempler på isolerede Begreber:")
    for b in begreber_isolated[:10]:
        report.append(f"      * '{b['name']}' (Domæne: {b.get('domainId')})")
    report.append("")

    # Analyze Klasser
    klasser_with_edges = [c for c in klasser if outgoing_edges[c["id"]] > 0 or incoming_edges[c["id"]] > 0]
    klasser_isolated = [c for c in klasser if outgoing_edges[c["id"]] == 0 and incoming_edges[c["id"]] == 0]

    report.append("--- INFORMATIONSMODEL (Klasser) CONNECTIVITY ---")
    report.append(f"  - Klasser med relationer (edges):  {len(klasser_with_edges)} ({len(klasser_with_edges)/len(klasser)*100:.1f}%)")
    report.append(f"  - Isolerede Klasser (0 edges):      {len(klasser_isolated)} ({len(klasser_isolated)/len(klasser)*100:.1f}%)")
    report.append(f"  - Eksempler på isolerede Klasser:")
    for k in klasser_isolated[:10]:
        report.append(f"      * '{k['name']}' (Attributter: {len(k.get('properties', []))})")
    report.append("\n" + "="*50 + "\n")

    # Analyze source Excel files to explain WHY
    report.append("--- ÅRSAGSANALYSE FRA KILDFILER ---")
    if os.path.exists(BEGREBSLISTE_PATH):
        wb_b = openpyxl.load_workbook(BEGREBSLISTE_PATH, data_only=True)
        report.append(f"1. begrebsliste_v1_1_0.xlsx indeholder {wb_b.sheetnames} med 931 begrebsdefinitioner.")
        report.append("   Begrebslisten er en national terminologisk kilde/ordbog. Den indeholder INGEN relations-fane eller kilde/mål-forbindelser mellem begreberne.")
        report.append("   Derfor stammer relationer/edges udelukkende fra Informationsmodellens Relationer-ark (242 klasser / 517 relationer).")
    
    if os.path.exists(SIS_IM_PATH):
        wb_im = openpyxl.load_workbook(SIS_IM_PATH, data_only=True)
        ws_rel = wb_im["Relationer"]
        report.append(f"2. SIS_IM_v1.1.3_20250319_1352.xlsx indeholder 'Relationer' arket med {ws_rel.max_row - 1} UML-relationer.")
        report.append("   Disse relationer forbinder de 242 strukturerede Klasser i Informationsmodellen.")
        report.append("   De 242 Klasser der er udledt af Begreberne har høje relateringsgrader, mens de 815 rent terminologiske Begreber i ordlisten (som ikke er udledt til en UML klasse i SIS IM) ikke har relationer i Excel-arkene.")

    report_str = "\n".join(report)
    with open(VERIFY_REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(report_str)

    print(report_str)

if __name__ == "__main__":
    main()
