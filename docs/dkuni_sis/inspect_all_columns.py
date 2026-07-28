import openpyxl

wb = openpyxl.load_workbook('/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/SIS_IM_v1.1.3_20250319_1352.xlsx', data_only=True)

report = []
for name in wb.sheetnames:
    ws = wb[name]
    report.append(f"==================================================")
    report.append(f"SHEET: {name} (rows: {ws.max_row}, cols: {ws.max_column})")
    report.append(f"==================================================")
    headers = [str(cell.value or '').strip() for cell in ws[1]]
    for idx, h in enumerate(headers, start=1):
        sample_vals = []
        for r in range(2, min(7, ws.max_row + 1)):
            v = ws.cell(r, idx).value
            if v is not None:
                sample_vals.append(str(v).strip().replace('\n', ' '))
        report.append(f"  Col {idx} [{h}]: samples -> {sample_vals[:3]}")
    report.append("\n")

with open('/home/rolfmadsen/Github/knowledgegraphstudio/docs/dkuni_sis/excel_full_structure.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(report))

print("Detailed report written to excel_full_structure.txt")
